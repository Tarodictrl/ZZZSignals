import os
import re
import subprocess
import tempfile
from datetime import datetime
from enum import Enum
from pathlib import Path
from re import Match
from time import sleep
from urllib.parse import parse_qs, urlsplit

import requests
from openpyxl import Workbook
from openpyxl.styles import Border, Font, PatternFill, Side

__version__ = "1.0.0"

RELEASE_URL = "https://api.github.com/repos/Tarodictrl/GenshinWishHistory/releases/latest"


class GachaType(Enum):
    STABLE = 1
    EVENT = 2
    WEAPON = 3
    BANBU = 5


class ZZZ:

    def __init__(self, url: str) -> None:
        self.__url, self.__params = self.__parse_url(url)

    @staticmethod
    def __parse_url(url: str):
        split = urlsplit(url)
        params = parse_qs(split.query)
        url = split.scheme + "://" + split.netloc + split.path
        return url, params

    def getBanner(self, real_gacha_type: GachaType | int) -> list:
        if isinstance(real_gacha_type, GachaType):
            real_gacha_type = real_gacha_type.value
        if isinstance(real_gacha_type, int) and real_gacha_type not in [1, 2, 3, 5]:
            raise ValueError("real_gacha_type must in: 1, 2, 3 or 5!")

        params = self.__params.copy()
        params.update(size=20, page=1, real_gacha_type=real_gacha_type)
        response = requests.get(self.__url, params=params)
        if response.status_code != 200 or response.json()["retcode"] != 0:
            return []

        response_json = response.json()
        signals = response_json["data"].get("list", [])

        while len(response_json["data"]["list"]) > 0:
            page = int(response_json["data"]["page"]) + 1
            end_id = signals[-1]["id"]
            params.update(page=page, end_id=end_id)

            response = requests.get(self.__url, params=params)
            if response.status_code != 200:
                break

            response_json = response.json()
            signals += response_json["data"].get("list", [])
            sleep(0.3)

        return signals


class Gacha:
    def __init__(self) -> None:
        self._api_url = "https://public-operation-nap-sg.hoyoverse.com"
        self._log_location = f"{os.environ['USERPROFILE']}\\AppData\\LocalLow\\miHoYo\\ZenlessZoneZero\\Player.log"

    @staticmethod
    def _getCacheUrl(match: Match[str]):
        game_dir = match.group()
        web_caches = sorted(Path(game_dir + "\\webCaches").iterdir(),
                            key=os.path.getmtime, reverse=True
                            )
        cache_file_path = f"{web_caches[0]}\\Cache\\Cache_Data\\data_2"
        temp_dir = tempfile.gettempdir()
        temp_path = os.path.join(temp_dir, 'temp_cache')
        os.system(f'echo.>{temp_path}')
        subprocess.run(['powershell', '-Command', f'copy "{cache_file_path}" "{temp_path}"'])
        with open(temp_path, encoding="utf-8", errors='ignore') as f:
            content = f.read()
        return content.split("1/0/")

    def loadLogs(self):
        if not os.path.exists(self._log_location):
            raise FileNotFoundError("Cannot find the log file! Make sure to open the wish history first!")
        with open(self._log_location) as f:
            return f.read()

    def loadCaches(self, logs: str) -> list:
        match = re.search(r"(.:/.+ZenlessZoneZero_Data)", logs, re.I)
        found = [x for x in self._getCacheUrl(match) if re.search("webview_gacha", x) is not None]
        return found

    def getLink(self, cache: str) -> str | None:
        link = re.findall(r"(https.+?end_id=)", cache)
        if not link:
            return
        test_result = self.testUrl(link[0])
        if test_result:
            return link[0]

    def testUrl(self, url: str) -> bool:
        try:
            response = requests.get(url=url)
            if response.status_code == 200:
                test_result = response.json()
                return test_result.get("retcode", -1) == 0
        except TimeoutError:
            print("Check link failed!")
        return False


class Saver:

    def __init__(self) -> None:
        self.wb = Workbook()
        self.wb.remove(self.wb["Sheet"])

    @property
    def purpleColor(self):
        return PatternFill(start_color='8B00FF',
                           end_color='8B00FF',
                           fill_type='solid'
                           )

    @property
    def goldColor(self):
        return PatternFill(start_color='FFD700',
                           end_color='FFD700',
                           fill_type='solid'
                           )

    @property
    def border(self):
        return Border(left=Side(style='thin'),
                      right=Side(style='thin'),
                      top=Side(style='thin'),
                      bottom=Side(style='thin'))

    def insert(self, sheet_name: str, data: list[dict]):
        if len(data) == 0:
            return
        keys = data[0].keys()
        if sheet_name not in self.wb.sheetnames:
            sheet = self.wb.create_sheet(sheet_name)
        else:
            sheet = self.wb["sheet_name"]
        for i, key in enumerate(keys, 0):
            place = f'{chr(65+i)}1'
            sheet[place].value = key.replace("_", " ").capitalize()
            sheet[place].font = Font(bold=True)

        for i, row in enumerate(data, 2):
            rank_type = row.get("rank", 0)
            for c, key in enumerate(keys, 0):
                place = f'{chr(65+c)}{i}'
                sheet[place].value = row[key]
                sheet[place].border = self.border
                if rank_type == '3':
                    sheet[place].fill = self.purpleColor
                if rank_type == '4':
                    sheet[place].fill = self.goldColor
        self.auto_width(sheet)

    def save(self, name: str | None = None) -> str:
        if not os.path.exists("signals"):
            os.mkdir("signals")
        if not name:
            name = "signals/signals_" + datetime.strftime(datetime.now(), "%Y_%m_%d_%H_%M") + ".xlsx"
        self.wb.save(name)
        return name

    def auto_width(self, sheet):
        for column in sheet.columns:
            max_length = 0
            column = [cell for cell in column]
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except Exception:
                    pass
            adjusted_width = (max_length + 6)
            sheet.column_dimensions[
                column[0].column_letter
            ].width = adjusted_width


def garant_counter(data: list[dict]):
    a_tier_counter = 0
    s_tier_counter = 0
    for row in reversed(data):
        a_tier_counter += 1
        s_tier_counter += 1
        rank_type = row["rank_type"]
        if rank_type == "3":
            row["count"] = a_tier_counter
            a_tier_counter = 0
        if rank_type == "4":
            row["count"] = s_tier_counter
            a_tier_counter = 0
            s_tier_counter = 0
    return data


def normalize_data(data: list[dict]):
    data = garant_counter(data)
    return [
        dict(name=x["name"], type=x["item_type"],
             rank=x["rank_type"], time=x["time"],
             count=x["count"]
             )
        for x in data
    ]


def printLogo():
    print("""
 ______   ______     ______     ______     _____     __     ______     ______   ______     __        
/\__  _\ /\  __ \   /\  == \   /\  __ \   /\  __-.  /\ \   /\  ___\   /\__  _\ /\  == \   /\ \       
\/_/\ \/ \ \  __ \  \ \  __<   \ \ \/\ \  \ \ \/\ \ \ \ \  \ \ \____  \/_/\ \/ \ \  __<   \ \ \____  
   \ \_\  \ \_\ \_\  \ \_\ \_\  \ \_____\  \ \____-  \ \_\  \ \_____\    \ \_\  \ \_\ \_\  \ \_____\ 
    \/_/   \/_/\/_/   \/_/ /_/   \/_____/   \/____/   \/_/   \/_____/     \/_/   \/_/ /_/   \/_____/                                                                                                                   
""")


def checkNeedUpdate() -> bool:
    response = requests.get(RELEASE_URL)
    if response.status_code == 200:
        latest_version = response.json().get("tag_name")
        if __version__ < latest_version:
            return True
    return False


if __name__ == "__main__":
    saver = Saver()
    gacha = Gacha()
    logs = gacha.loadLogs()
    caches = gacha.loadCaches(logs)
    for i in range(len(caches) - 1, -1, -1):
        os.system("cls")
        printLogo()
        print(f"Checking link: {i}\n")
        link = gacha.getLink(caches[i])
        if link:
            print(link)
            zzz = ZZZ(link)
            saver.insert("Event", normalize_data(zzz.getBanner(2)))
            saver.insert("Stable", normalize_data(zzz.getBanner(1)))
            saver.insert("Banbu", normalize_data(zzz.getBanner(5)))
            saver.save()
            print("\033[92mDone!\x1b[0m")
            print("\033[92mFile saved as:\x1b[0m", saver.save())
            flag = True
            break
        sleep(1)
    for i in range(9, 0, -1):
        print(f"Window will close after: {i} s", end="\r", flush=True)
        sleep(1)
